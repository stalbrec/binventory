from django.views import generic
from .models import Box, Item
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .utils import get_qr_code_buffer,generate_location_svg
import base64
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl import Workbook, load_workbook
from django import forms
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView
from django.contrib.auth.forms import AuthenticationForm

# Create your views here.
import logging
class CustomLoginForm(AuthenticationForm):
    username = forms.CharField(
        widget=forms.TextInput(attrs={'class':'input input-bordered w-full', 'placeholder':'Username'})
    )
    password = forms.CharField(
        widget=forms.PasswordInput(attrs={
            'class':'input input-bordered w-full', 'placeholder':'Password'
        })
    )

class CustomLoginView(LoginView):
    form_class = CustomLoginForm
    template_name = "registration/login.html"
    redirect_authenticated_user=True
    next_page="/"

class ExcelImportForm(forms.Form):
    # title = forms.CharField(max_length=50)
    mode = forms.ChoiceField(choices=[
        ("append", "append"), 
        # ("replace", "replace")
        ])
    file = forms.FileField()

@login_required
def import_excel(request):
    if request.method == "POST":
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            buffer = BytesIO(request.FILES["file"].read())
            wb = load_workbook(buffer)
            ws = wb.active
            # import_mode = form.cleaned_data["mode"]
            for irow in range(2, ws.max_row + 1):
                row = ws[irow]
                item_id, item_name, box_name, box_location = row
                box_query = Box.objects.filter(name=box_name.value)
                if box_query.exists():
                    box = box_query[0]
                else:
                    box = Box.objects.create(
                        name=box_name.value, location=box_location.value
                    )
                Item.objects.create(name=item_name.value, box=box)
        return redirect("inventory:index")
    else:
        form = ExcelImportForm()
    return render(request, "inventory/import.html", {"form": form})

class IndexView(LoginRequiredMixin, generic.ListView):
    template_name = "inventory/index.html"
    context_object_name = "full_inventory"

    def get_queryset(self):
        return Item.objects.order_by("name")


class BoxAwareDetailView(generic.DetailView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["available_boxes"] = Box.objects.all()
        return context


class ItemView(LoginRequiredMixin, BoxAwareDetailView):
    model = Item
    template_name = "inventory/item.html"

    def post(self, request, *args, **kwargs):
        pk = kwargs.get("pk")
        if request.method == "POST":
            obj = self.model.objects.get(id=pk)
            new_box_id = request.POST.get("new_box")
            new_box = Box.objects.get(id=new_box_id)
            if new_box:
                obj.box = new_box
            obj.save()
        return redirect("inventory:item", pk)
    
    def get_context_data(self, **kwargs):
        context =  super().get_context_data(**kwargs)
        svg_str = generate_location_svg(self.object.box.location,"Hobbyraum")
        if svg_str == "":
            svg_str=(
                '<svg xmlns="http://www.w3.org/2000/svg" width="200" height="200" viewBox="0 0 200 200">'
                '<rect width="100%" height="100%" rx="16" ry="16" fill="#f5f5f5" stroke="#ccc" stroke-width="2"/>'
                '<path d="M40 130 L70 100 L100 130 L130 100 L160 130" stroke="#bbb" stroke-width="4" fill="none" stroke-linecap="round" stroke-linejoin="round"/>'
                '<circle cx="60" cy="70" r="8" fill="#bbb"/>'
                '<line x1="40" y1="40" x2="160" y2="160" stroke="#e0e0e0" stroke-width="3"/>'
                '<line x1="160" y1="40" x2="40" y2="160" stroke="#e0e0e0" stroke-width="3"/>'
                '</svg>'
            )
        img_base64 = base64.b64encode(svg_str.encode()).decode()

        context["location_image_data"] = f"data:image/svg+xml;base64,{img_base64}"
        return context

class BoxView(LoginRequiredMixin, BoxAwareDetailView):
    model = Box
    template_name = "inventory/box.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        full_url = self.request.build_absolute_uri(self.object.get_absolute_url())
        context["full_url"] = full_url

        buffer = get_qr_code_buffer(full_url)
        img_str = base64.b64encode(buffer.getvalue())
        context["qr_image_data"] = f"data:image/svg+xml;base64,{img_str.decode()}"
        logging.error(img_str)
        return context

@login_required
def download_all_box_qr_codes(request):
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "a", ZIP_DEFLATED) as zip_file:
        for box in Box.objects.order_by("id"):
            box_full_url = request.build_absolute_uri(box.get_absolute_url())
            zip_file.writestr(
                f"{box.name}.svg", get_qr_code_buffer(box_full_url).getvalue()
            )
    zip_buffer.seek(0)

    response = HttpResponse(
        zip_buffer.read(),
        content_type="application/zip",
    )
    downlaod_fname = "box_qr_codes.zip"
    response["Content-Disposition"] = f'attachment; filename="{downlaod_fname}"'
    return response


@login_required
def export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    columns = ["item-id", "item-name", "box-name", "box-location"]
    ws.append(columns)

    for item in Item.objects.order_by("id"):
        ws.append([item.id, item.name, item.box.name, item.box.location])
    xlsx_buffer = BytesIO()
    wb.save(xlsx_buffer)

    xlsx_buffer.seek(0)

    response = HttpResponse(
        xlsx_buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.speadsheetml.sheet",
    )
    downlaod_fname = "inventory.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{downlaod_fname}"'
    return response


@login_required
def new_item(request):
    if request.method == "POST":
        item_name = request.POST.get("name")
        box_id = request.POST.get("box")
        Item.objects.create(name=item_name, box=Box.objects.get(id=box_id))
        return redirect("inventory:index")
    return render(
        request, "inventory/new.html", dict(available_boxes=Box.objects.all())
    )
